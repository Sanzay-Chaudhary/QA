from openpyxl import load_workbook


def read_excel(file_path, LoginData):

    workbook = load_workbook(file_path)

    sheet = workbook[LoginData]

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        data.append(row)

    workbook.close()

    return data
